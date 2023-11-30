import os
import openpyxl
import json

root_path = os.path.abspath('.')
xlsx_skill_file = root_path + "\\xlsx\\" + "Skill_EffectConfig.xlsx"
xlsx_magic_file = root_path + "\\xlsx\\" + "Magic_EffectConfig.xlsx"
json_skill_file = root_path + "\\Setting\\"  + "Skill.json"
json_magic_file = root_path + "\\Setting\\"  + "Magic.json"


def xlsx_convert(rf_path, sf_path):
    data_xlsx = openpyxl.load_workbook(rf_path, data_only=False).active

    excel_data_dict = {}
    col_name_list = []
    col_val_type_list = []
    data_dict = {
        "root": {
            "ID" : [],
            "ver" : 0
        }
    }
    
    ver_num = data_xlsx.cell(1, 2)

    # 遍历保存字段名
    for col in range(1, data_xlsx.max_column + 1):
        cell = data_xlsx.cell(4, col)
        v = cell.value
        col_name_list.append(str(v))

    # 遍历保存数据类型
    for col in range(1, data_xlsx.max_column + 1):
        cell = data_xlsx.cell(5, col)
        v = cell.value
        col_val_type_list.append(str(v))

    # 剔除表头、字段名和字段类型所在行 从第四行开始遍历 构造行数据
    for row in range(6, data_xlsx.max_row + 1):
        row_data_dict = {}
        cell_data_dict = {}
        # 保存数据索引 默认第一列为id
        cell_id = data_xlsx.cell(row, 1)
        
        # 检查id的唯一性
        if cell_id.value in excel_data_dict:
            print('[warning] duplicated data id: "%d", all previous value will be ignored!~' % (cell_id.value)) 
 
        # 保存每一行的所有数据
        for col in range(1, data_xlsx.max_column + 1):
            cell = data_xlsx.cell(row, col)
            key = str(col_name_list[col - 1])
            cell_val_type = str(col_val_type_list[col - 1])
            
            # 根据字段类型去调整数值 如果为空值 依据字段类型 填上默认值
            if cell.value == None:
                    continue
            elif cell_val_type == 'string':
                v = str(cell.value)
            elif cell_val_type == 'int':
                v = int(cell.value)
            elif cell_val_type == 'float':
                v = float(cell.value)
            elif cell_val_type == 'dict':
                v = json.loads(str(cell.value))
            # 加入列表
            row_data_dict[key] = v
        # 保存id 和 row data
        cell_data_dict["d"] = cell_id.value
        cell_data_dict["h"] = row_data_dict
        data_dict["root"]["ID"].append(cell_data_dict)
        
    data_dict["root"]["ver"] = ver_num.value
    
    with open(sf_path, "w", encoding='utf-8') as f:
        json.dump(data_dict, f, indent=3, sort_keys=True, ensure_ascii=False)  # 写为多行


if __name__ == '__main__':
    print("开始导出：Magic.json & Skill.json")
    try:
        xlsx_convert(xlsx_magic_file, json_magic_file)
        xlsx_convert(xlsx_skill_file, json_skill_file)
    except Exception as e:
        print(e)
    print("Magic.json & Skill.json 导出完成")
    os.system("pause")